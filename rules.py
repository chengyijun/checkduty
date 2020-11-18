# -*- coding:utf-8 -*-
"""
@author: chengyijun
@contact: cyjmmy@foxmail.com
@file: rules.py
@time: 2020/11/16 14:46
@desc:
"""
import json
import re
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def main():
    Rule()


class Rule:
    def __init__(self):
        self.write_result()

    def write_result(self):
        # 读取数据
        with open('./datas/results.json', 'r', encoding='utf-8') as f:
            results = json.load(f)

        datas = self.__get_results_by_rules(results)

        wb = load_workbook(r'./template.xlsx')
        ws = wb.active

        for data in datas:
            for item in data.items():
                # {'D5': '张宽', 'T5': 25, 'U5': 375, 'V5': 25, 'W5': 217.39, 'X5': 7, 'Y5': 23, 'Z5': 23}
                if 'AA' in str(item[0]):
                    # 备注信息靠左对齐
                    ws[item[0]].alignment = Alignment(horizontal='left', vertical='center')
                ws[item[0]] = str(item[1])

        wb.save(r'./static/考勤结果.xlsx')
        # print('处理完毕')

    def __get_results_by_rules(self, results: list) -> list:
        """
        根据考勤规则 计算个人的结果
        :param result:
        :return:
        """
        datas = []
        for row, result in enumerate(results, start=5):
            # 备注信息
            beizhus = []

            # 工作日加班  餐补+1
            cb1 = self.deal_jzrjb(result)
            # 公休日加班 餐补+1 交补+1
            cb2, jb1 = self.deal_jrjb(result)
            jb_cb_count = cb1 + cb2
            jb_jb_count = jb1

            # 提取入职信息 加入备注
            self.deal_ruzhi(beizhus, result)
            # 提取缺卡信息
            self.deal_queka(beizhus, result)
            # 提取补卡信息
            self.deal_buka(beizhus, result)

            # 处理出差  餐补-1 交补-1
            cc_count = self.deal_cc(beizhus, result)

            jb_jb_count -= cc_count
            jb_cb_count -= cc_count

            beizhu_str = ' '.join(beizhus) if beizhus else ''
            print(beizhu_str)

            # 写结果到模版文件
            data = {
                f'D{row}': result['姓名'],
                f'T{row}': result['实际出勤天数'] + jb_cb_count,
                f'U{row}': round((result['实际出勤天数'] + jb_cb_count) * 15, 2),
                f'V{row}': result['实际出勤天数'] + jb_jb_count,
                f'W{row}': round((result['实际出勤天数'] + jb_jb_count) * 200 / 23, 2),
                f'X{row}': result['公休天数'],
                f'Y{row}': result['应该出勤天数'],
                f'Z{row}': result['实际出勤天数'],
                f'AA{row}': beizhu_str
            }
            datas.append(data)

        return datas

    def deal_buka(self, beizhus: list, result: dict) -> None:
        """
        处理补卡
        :param beizhus:
        :param result:
        :return:
        """
        temps = []
        if result['补卡']:
            for bk in result['补卡']:
                temps.append(bk.split(' ')[0][3:])
            beizhus.append('补卡：' + ' '.join(temps))

    def deal_queka(self, beizhus: list, result: dict) -> None:
        """
        处理缺卡
        :param beizhus:
        :param result:
        :return:
        """
        temps = []
        if result['缺卡']:
            for qk in result['缺卡']:
                temps.append(qk.split(' ')[0][3:])
            beizhus.append('缺卡：' + ' '.join(temps))

    def deal_ruzhi(self, beizhus: list, result: dict):
        """
        如果有入职信息，则加入备注列表
        :param beizhus:
        :param result:
        :return:
        """
        if result['入职信息'][0]:
            beizhu = '入职：' + result['入职信息'][2]
            beizhus.append(beizhu)

    def deal_jrjb(self, result: dict) -> Tuple[int, int]:
        """
        处理假日加班的餐补交补
        :param result:
        :return:
        """
        jb_cb_count = 0
        jb_jb_count = 0
        jr_jbs = result['假日加班']
        if jr_jbs:
            for jr_jb in jr_jbs:
                gp = re.search(r'\s(\d+(\.\d+)?)小时', jr_jb)
                if gp:
                    if float(gp.group(1)) > 2.0:
                        # 有效加班
                        jb_cb_count += 1
                        jb_jb_count += 1
        return jb_cb_count, jb_jb_count

    def deal_jzrjb(self, result: dict) -> int:
        """
        处理工作日加班的餐补交补
        :param result:
        :return:
        """
        jb_cb_count = 0
        jzr_jbs = result['工作日加班']
        if jzr_jbs:
            for jzr_jb in jzr_jbs:
                gp = re.search(r'\s(\d+(\.\d+)?)小时', jzr_jb)
                if gp:
                    if float(gp.group(1)) > 2.0:
                        # 有效加班
                        jb_cb_count += 1
        return jb_cb_count

    def deal_cc(self, beizhus: list, result: dict) -> int:
        """
        处理出差
        :param beizhus:
        :param result:
        :return:
        """
        # "出差09-03 08:30到09-03 18:00 1天"
        cc_count = 0
        ccs = result['出差']
        if ccs:
            for cc in ccs:
                try:
                    gp = re.search(r'\s(\d+)天', cc)
                except Exception as e:
                    print(e)
                else:
                    cc_count += int(gp.group(1))
            beizhus.append(f'出差：{cc_count}天')
        return cc_count


if __name__ == '__main__':
    main()
