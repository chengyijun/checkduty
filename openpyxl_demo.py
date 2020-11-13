# -*- coding:utf-8 -*-
"""
@author: chengyijun
@contact: cyjmmy@foxmail.com
@file: openpyxl_demo.py
@time: 2020/11/11 10:16
@desc:
"""


def main():
    from openpyxl import load_workbook
    wb = load_workbook(r'./template.xlsx')
    ws = wb.active

    # print(ws.title)
    ws['Z5'] = 23
    wb.save(r'./template.xlsx')


if __name__ == '__main__':
    main()
