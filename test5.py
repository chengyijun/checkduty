# -*- coding:utf-8 -*-
"""
@author: chengyijun
@contact: cyjmmy@foxmail.com
@file: test5.py
@time: 2020/11/10 13:27
@desc:
"""
import json
import re
import time
from pprint import pprint
from typing import List, Tuple

import pandas as pd
from openpyxl import load_workbook
from pandas import Index, DataFrame

from utils import get_days


class Duty:
    # 1. 获取到了需要统计的人员名单 （排除不在考勤组的，排除销售的）
    df1 = pd.read_excel('./datas/1.xlsx', skiprows=2)
    # 1. 读取考勤数据 每日统计
    df2 = pd.read_excel('./datas/1.xlsx', skiprows=2, sheet_name='每日统计')
    # 获取当前年份
    year = time.strftime("%Y", time.localtime())

    def __init__(self):
        self.holidays = self.get_holidays()
        self.peoples = self.__get_duty_people_range()
        self.current_month = self.__get_current_month()

    @staticmethod
    def get_holidays() -> list:
        """
        获取全年假日日期列表
        :return:
        """
        with open('./holidays.json', 'r') as f:
            dates = f.read()
        return json.loads(dates)

    @staticmethod
    def __get_current_month() -> int:
        """
        获取当前被统计的是哪个月份
        :return:
        """
        wb = load_workbook(r'./datas/1.xlsx')
        ws = wb["月度汇总"]
        return int(str(ws['A1'].value)[-5:-3])

    def __get_duty_people_range(self) -> Index:
        """
        获取到了需要统计的人员名单 （排除不在考勤组的，排除销售的）
        :return:
        """
        self.df1.set_index('姓名', inplace=True)
        self.df1.drop(self.df1.index[0], inplace=True)
        condition1 = self.df1['考勤组'].str.contains('未加入考勤组')
        condition2 = self.df1['部门'].str.contains('营销中心')
        condition = ~(condition1 | condition2)
        df2 = self.df1[condition]
        df3 = df2.loc[:, '1':].T
        # 获取到了需要统计的人员名单 （排除不在考勤组的，排除销售的）
        return df3.columns

    def __get_duty_data(self, name: str) -> dict:
        """
        获取每个人的考勤信息
        :param name:
        :return:
        """
        cond = self.df2['姓名'] == name
        df2 = self.df2[cond]

        # 判断是否本月入职，实际工作天数是多少，入职日期是当月多少号
        ruzhi_infos = self.__get_ruzhi_infos(df2)
        # 获取迟到信息
        chidao_dict = self.__get_chidao_infos(df2)
        # 获取补卡、缺卡信息
        bukas, quekas = self.__get_buka_queka_infos(df2, ruzhi_infos)
        # 获取关联审批单
        new_spds = self.__get_spds(df2)

        txs = self.__get_category_infos(new_spds, '调休')
        sjs = self.__get_category_infos(new_spds, '事假')
        bjs = self.__get_category_infos(new_spds, '病假')
        njs = self.__get_category_infos(new_spds, '年假')
        hjs = self.__get_category_infos(new_spds, '婚假')
        ssjs = self.__get_category_infos(new_spds, '丧假')
        jbs = self.__get_category_infos(new_spds, '加班')
        ccs = self.__get_category_infos(new_spds, '出差')

        # 获取工作日加班和假日加班信息
        gzr_jbs, jr_jbs = self.__get_jbs_by_jiari_gongzuori(jbs)
        # 获取本月出差信息 排除跨月出差的天数
        real_ccs = self.__get_real_chuchai_infos(ccs)

        data = {
            '姓名': name,
            '迟到': chidao_dict,
            '补卡': bukas,
            '缺卡': quekas,
            '事假': sjs,
            '调休': txs,
            '病假': bjs,
            '年假': njs,
            '婚假': hjs,
            '丧假': ssjs,
            '工作日加班': gzr_jbs,
            '假日加班': jr_jbs,
            '出差': real_ccs,
            '工作天数': ruzhi_infos[1],
            '入职信息': ruzhi_infos

        }
        return data

    def __get_real_chuchai_infos(self, ccs: List[str]) -> List[str]:
        """
        获取本月的出差信息 排除跨月的天数
        :param ccs:
        :return:
        """
        real_ccs = []
        for cc in ccs:
            gp = re.search(r'出差(\d\d-\d\d) .*?到(\d\d-\d\d) .*? (\d+)天', cc)
            if int(gp.group(1).split('-')[0]) != int(gp.group(2).split('-')[0]):
                # 跨了月份  出差天数就是 int(gp.group(3))-n
                sd = f"{self.year}{gp.group(1).replace('-', '')}"
                ed = f"{self.year}{gp.group(2).replace('-', '')}"
                n = get_days(sd, ed, self.current_month)
                cc = re.sub(r' \d+天', f' {str(int(gp.group(3)) - n)}天', cc)
            real_ccs.append(cc)
        return real_ccs

    def __get_jbs_by_jiari_gongzuori(self, jbs: List[str]) -> Tuple[List[str], List[str]]:
        """
        获取工作日加班和工作日加班信息
        :param jbs:
        :return:
        """
        gzr_jbs = []
        jr_jbs = []
        for jb in jbs:
            gp = re.search(r'加班(\d\d-\d\d)\s.*?', jb)

            date = f"{self.year}{gp.group(1).replace('-', '')}"

            if date in self.holidays:
                jr_jbs.append(jb)
            else:
                gzr_jbs.append(jb)
        return gzr_jbs, jr_jbs

    @staticmethod
    def __get_category_infos(new_spds: List[str], prefix: str) -> List[str]:
        """
        根据不同的前缀 从审批单中获取对应信息
        :param new_spds:
        :param prefix:
        :return:
        """
        temps = []
        for new_spd in new_spds:
            if prefix in new_spd:
                temps.append(new_spd)
        return temps

    @staticmethod
    def __get_spds(df2: DataFrame) -> List[str]:
        """
        获取关联审批单
        :param df2:
        :return:
        """
        df3 = df2[df2['关联的审批单'].notna()]
        spds = list(df3['关联的审批单'].values)
        new_spds = []
        for spd in spds:
            if '\n' in spd:
                temps = spd.split('\n')
                new_spds.extend(temps)
            else:
                new_spds.append(spd)
        new_spds = list(set(new_spds))
        return new_spds

    @staticmethod
    def __get_buka_queka_infos(df2: DataFrame, ruzhi_infos: Tuple[bool, int, str]) -> Tuple[List[str], List[str]]:
        """
        获取补卡、缺卡信息
        :param df2:
        :return:
        """
        df5 = df2[(df2['Unnamed: 9'].notna()) & (df2['Unnamed: 11'].notna())]
        df6 = df5[(df5['Unnamed: 11'].str.contains('补卡')) | (df5['Unnamed: 9'].str.contains('补卡'))]
        df7 = df5[(df5['Unnamed: 11'].str.contains('缺卡')) | (df5['Unnamed: 9'].str.contains('缺卡'))]
        bks = list(df6['日期'].to_dict().values())
        qks = list(df7['日期'].to_dict().values())
        real_bks = []
        real_qks = []
        if ruzhi_infos[0]:
            # '20-09-16 星期三'  '20200916'
            target = ruzhi_infos[2][4:]
            temp = f'{target[:2]}-{target[2:]}'
            for bk in bks:
                if temp not in bk:
                    real_bks.append(bk)
            for qk in qks:
                if temp not in qk:
                    real_qks.append(qk)
            return real_bks, real_qks
        return bks, qks

    @staticmethod
    def __get_chidao_infos(df2: DataFrame) -> dict:
        """
        获取迟到信息
        :param df2:
        :return:
        """
        df4 = df2[(df2['Unnamed: 9'].notna()) & (df2['Unnamed: 9'].str.contains('迟到'))]
        return df4[['日期', 'Unnamed: 9']].set_index('日期')['Unnamed: 9'].to_dict()

    def __get_ruzhi_infos(self, df2: DataFrame) -> Tuple[bool, int, str]:
        """
        判断并获取入职信息
        :param df2:
        :return:
        """
        real_work_days_count = 23
        is_ruzhi_current_month = False
        no_offer_days = list(df2[df2['班次'] == '不在考勤组']['日期'].values)
        if len(no_offer_days) != 0:
            is_ruzhi_current_month = True
            start = self.year + no_offer_days[-1].split(' ')[0][2:]
            end = f'{self.year}-{str(self.current_month + 1)}-01' if self.current_month + 1 >= 10 else f'{self.year}-0{str(self.current_month + 1)}-01 '
            days = pd.date_range(start=start, end=end)
            targets = [pd.Timestamp(x).strftime("%Y%m%d") for x in days.values]
            targets.pop()
            targets.pop(0)
            real_work_days = []
            for target in targets:
                if target not in self.holidays:
                    real_work_days.append(target)
            real_work_days_count = len(real_work_days)
            return is_ruzhi_current_month, real_work_days_count, real_work_days[0]
        return is_ruzhi_current_month, real_work_days_count, ''

    def get_duty_datas(self) -> List[dict]:
        """
        获取所有人考勤汇总信息
        :return:
        """
        datas = []
        for people in self.peoples:
            datas.append(self.__get_duty_data(people))
        return datas


def main():
    duty = Duty()
    pprint(duty.get_duty_datas())


if __name__ == '__main__':
    main()
